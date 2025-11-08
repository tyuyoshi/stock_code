"""Export service layer for data export functionality"""

import io
import csv
import uuid
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional, BinaryIO
from sqlalchemy.orm import Session
from fastapi import HTTPException
from fastapi.responses import StreamingResponse

from models.company import Company
from models.financial import FinancialStatement, FinancialIndicator, StockPrice
from schemas.export import (
    ExportFormat,
    ExportDataType,
    CompaniesExportRequest,
    ScreeningExportRequest,
    ComparisonExportRequest,
    FinancialDataExportRequest,
    ExportTemplate,
)
from services.company_service import CompanyService
from services.screening_service import ScreeningService
from services.compare_service import CompareService
from schemas.screening import ScreeningRequest
from schemas.compare import CompareRequest


class ExportService:
    """Service class for data export functionality"""

    @staticmethod
    def export_companies(
        db: Session, request: CompaniesExportRequest
    ) -> StreamingResponse:
        """Export companies data"""

        # Get companies data
        if request.company_ids:
            companies = (
                db.query(Company).filter(Company.id.in_(request.company_ids)).all()
            )
        else:
            companies = db.query(Company).all()

        # Get financial indicators if requested
        indicators_map = {}
        if request.include_indicators:
            for company in companies:
                indicator = (
                    db.query(FinancialIndicator)
                    .filter(FinancialIndicator.company_id == company.id)
                    .order_by(FinancialIndicator.date.desc())
                    .first()
                )
                indicators_map[company.id] = indicator

        # Determine fields to include
        if request.fields:
            fields = request.fields
        else:
            fields = ExportService._get_default_company_fields(
                request.include_indicators
            )

        # Generate export data
        if request.format == ExportFormat.CSV:
            return ExportService._export_companies_csv(
                companies, indicators_map, fields, request
            )
        elif request.format == ExportFormat.EXCEL:
            return ExportService._export_companies_excel(
                companies, indicators_map, fields, request
            )
        else:
            raise HTTPException(status_code=400, detail="Unsupported export format")

    @staticmethod
    def _export_companies_csv(
        companies: List[Company],
        indicators_map: Dict[int, FinancialIndicator],
        fields: List[str],
        request: CompaniesExportRequest,
    ) -> StreamingResponse:
        """Export companies data as CSV"""

        output = io.StringIO()
        writer = csv.writer(output)

        # Write headers
        if request.include_headers:
            headers = ExportService._get_field_labels(fields)
            writer.writerow(headers)

        # Write data rows
        for company in companies:
            row = []
            indicator = indicators_map.get(company.id)

            for field in fields:
                value = ExportService._get_field_value(company, indicator, field)
                row.append(value)

            writer.writerow(row)

        output.seek(0)

        filename = (
            request.filename or f"companies_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        )

        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
        )

    @staticmethod
    def _export_companies_excel(
        companies: List[Company],
        indicators_map: Dict[int, FinancialIndicator],
        fields: List[str],
        request: CompaniesExportRequest,
    ) -> StreamingResponse:
        """Export companies data as Excel"""

        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise HTTPException(
                status_code=500, detail="Excel export requires openpyxl package"
            )

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Companies"

        # Write headers
        if request.include_headers:
            headers = ExportService._get_field_labels(fields)
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)

        # Write data rows
        start_row = 2 if request.include_headers else 1
        for row_idx, company in enumerate(companies, start_row):
            indicator = indicators_map.get(company.id)

            for col_idx, field in enumerate(fields, 1):
                value = ExportService._get_field_value(company, indicator, field)
                worksheet.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = (
            request.filename or f"companies_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        )

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
        )

    @staticmethod
    def _get_default_company_fields(include_indicators: bool) -> List[str]:
        """Get default fields for company export"""
        fields = [
            "ticker_symbol",
            "company_name_jp",
            "company_name_en",
            "market_division",
            "industry_name",
            "market_cap",
            "employee_count",
            "listing_date",
        ]

        if include_indicators:
            fields.extend(
                [
                    "roe",
                    "roa",
                    "operating_margin",
                    "debt_to_equity",
                    "current_ratio",
                    "per",
                    "pbr",
                    "dividend_yield",
                ]
            )

        return fields

    @staticmethod
    def _get_field_labels(fields: List[str]) -> List[str]:
        """Get human-readable labels for fields"""
        label_map = {
            "ticker_symbol": "ティッカー",
            "company_name_jp": "会社名",
            "company_name_en": "会社名（英語）",
            "market_division": "市場区分",
            "industry_name": "業種",
            "market_cap": "時価総額（百万円）",
            "employee_count": "従業員数",
            "listing_date": "上場日",
            "roe": "ROE（%）",
            "roa": "ROA（%）",
            "operating_margin": "営業利益率（%）",
            "equity_ratio": "自己資本比率（%）",
            "current_ratio": "流動比率（%）",
            "per": "PER（倍）",
            "pbr": "PBR（倍）",
            "dividend_yield": "配当利回り（%）",
            "debt_to_equity": "負債資本倍率",
        }

        return [label_map.get(field, field) for field in fields]

    @staticmethod
    def _get_field_value(
        company: Company, indicator: Optional[FinancialIndicator], field: str
    ) -> Any:
        """Get field value from company or indicator"""

        # Company fields
        if hasattr(company, field):
            value = getattr(company, field)
            if field == "listing_date" and value:
                return value.strftime("%Y-%m-%d")
            return value

        # Indicator fields
        if indicator and hasattr(indicator, field):
            return getattr(indicator, field)

        return None

    @staticmethod
    def get_export_templates() -> List[ExportTemplate]:
        """Get predefined export templates"""
        templates = [
            ExportTemplate(
                id="basic_companies",
                name="基本企業情報",
                description="ティッカー、会社名、市場区分、業種等の基本情報",
                data_type=ExportDataType.COMPANIES,
                format=ExportFormat.CSV,
                fields=[
                    "ticker_symbol",
                    "company_name_jp",
                    "market_division",
                    "industry_name",
                ],
                category="basic",
            ),
            ExportTemplate(
                id="financial_overview",
                name="財務概要",
                description="基本企業情報 + 主要財務指標",
                data_type=ExportDataType.COMPANIES,
                format=ExportFormat.EXCEL,
                fields=[
                    "ticker_symbol",
                    "company_name_jp",
                    "market_division",
                    "market_cap",
                    "roe",
                    "roa",
                    "operating_margin",
                    "debt_to_equity",
                    "per",
                    "pbr",
                ],
                category="financial",
            ),
            ExportTemplate(
                id="investment_analysis",
                name="投資分析用データ",
                description="投資判断に必要な主要指標一覧",
                data_type=ExportDataType.COMPANIES,
                format=ExportFormat.EXCEL,
                fields=[
                    "ticker_symbol",
                    "company_name_jp",
                    "market_cap",
                    "per",
                    "pbr",
                    "dividend_yield",
                    "roe",
                    "debt_to_equity",
                    "revenue_growth_yoy",
                    "earnings_growth_yoy",
                ],
                category="investment",
            ),
            ExportTemplate(
                id="screening_results",
                name="スクリーニング結果",
                description="スクリーニング結果のエクスポート用テンプレート",
                data_type=ExportDataType.SCREENING_RESULTS,
                format=ExportFormat.CSV,
                fields=[
                    "ticker_symbol",
                    "company_name_jp",
                    "market_division",
                    "roe",
                    "per",
                    "dividend_yield",
                ],
                category="screening",
            ),
            ExportTemplate(
                id="comparison_report",
                name="企業比較レポート",
                description="複数企業の比較分析レポート",
                data_type=ExportDataType.COMPARISON,
                format=ExportFormat.EXCEL,
                fields=[
                    "ticker_symbol",
                    "company_name_jp",
                    "roe",
                    "roa",
                    "per",
                    "pbr",
                    "debt_to_equity",
                    "current_ratio",
                ],
                category="comparison",
            ),
        ]
        return templates

    @staticmethod
    def export_screening_results(
        db: Session, request: ScreeningExportRequest
    ) -> StreamingResponse:
        """Export screening results"""

        # Convert export request to screening request
        screening_request = ScreeningRequest(
            filters=request.filters,
            sort=request.sort,
            page=1,
            size=min(request.max_rows, 10000),
            include_indicators=request.include_indicators,
        )

        # Execute screening
        results, total, _ = ScreeningService.execute_screening(db, screening_request)

        # Limit results to max_rows
        if len(results) > request.max_rows:
            results = results[: request.max_rows]

        # Determine fields
        if hasattr(request, "fields") and request.fields:
            fields = request.fields
        else:
            fields = ExportService._get_default_company_fields(
                request.include_indicators
            )

        # Generate export
        if request.format == ExportFormat.CSV:
            return ExportService._export_screening_csv(results, fields, request)
        elif request.format == ExportFormat.EXCEL:
            return ExportService._export_screening_excel(results, fields, request)
        else:
            raise HTTPException(status_code=400, detail="Unsupported export format")

    @staticmethod
    def _export_screening_csv(
        results: List, fields: List[str], request: ScreeningExportRequest
    ) -> StreamingResponse:
        """Export screening results as CSV"""

        output = io.StringIO()
        writer = csv.writer(output)

        # Write filter info as comment
        writer.writerow(
            [f"# Screening Filters: {len(request.filters)} filters applied"]
        )
        for filter_item in request.filters:
            writer.writerow(
                [
                    f"# {filter_item.get('field')} {filter_item.get('operator')} {filter_item.get('value')}"
                ]
            )
        writer.writerow([])  # Empty line

        # Write headers
        if request.include_headers:
            headers = ExportService._get_field_labels(fields)
            writer.writerow(headers)

        # Write data rows
        for result in results:
            row = []
            for field in fields:
                # Check if field is in indicators dict first
                if hasattr(result, "indicators") and result.indicators and field in result.indicators:
                    value = result.indicators[field]
                else:
                    value = getattr(result, field, None)
                if hasattr(value, "strftime"):
                    value = value.strftime("%Y-%m-%d")
                row.append(value)
            writer.writerow(row)

        output.seek(0)

        filename = (
            request.filename or f"screening_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        )

        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
        )

    @staticmethod
    def _export_screening_excel(
        results: List, fields: List[str], request: ScreeningExportRequest
    ) -> StreamingResponse:
        """Export screening results as Excel"""

        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font
        except ImportError:
            raise HTTPException(
                status_code=500, detail="Excel export requires openpyxl package"
            )

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Screening Results"

        # Write filter info
        row_idx = 1
        worksheet.cell(row=row_idx, column=1, value="Screening Filters")
        row_idx += 1
        for filter_item in request.filters:
            filter_text = f"{filter_item.get('field')} {filter_item.get('operator')} {filter_item.get('value')}"
            worksheet.cell(row=row_idx, column=1, value=filter_text)
            row_idx += 1
        row_idx += 1  # Empty row

        # Write headers
        if request.include_headers:
            headers = ExportService._get_field_labels(fields)
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=row_idx, column=col, value=header)
            row_idx += 1

        # Write data rows
        for result in results:
            for col_idx, field in enumerate(fields, 1):
                # Check if field is in indicators dict first
                if hasattr(result, "indicators") and result.indicators and field in result.indicators:
                    value = result.indicators[field]
                else:
                    value = getattr(result, field, None)
                if hasattr(value, "strftime"):
                    value = value.strftime("%Y-%m-%d")
                worksheet.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = (
            request.filename or f"screening_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        )

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
        )

    @staticmethod
    def export_comparison(
        db: Session, request: ComparisonExportRequest
    ) -> StreamingResponse:
        """Export company comparison results"""

        # Convert export request to compare request
        from schemas.compare import CompareRequest

        compare_request = CompareRequest(
            company_ids=request.company_ids,
            metrics=request.metrics,
            include_rankings=request.include_rankings,
        )

        # Get comparison data
        companies, summary = CompareService.compare_companies(db, compare_request)

        # Only Excel format for comparison (better for side-by-side comparison)
        if request.format == ExportFormat.EXCEL:
            return ExportService._export_comparison_excel(companies, summary, request)
        else:
            raise HTTPException(
                status_code=400,
                detail="Only Excel format is supported for comparison export",
            )

    @staticmethod
    def _export_comparison_excel(
        companies: List, summary: Any, request: ComparisonExportRequest
    ) -> StreamingResponse:
        """Export comparison results as Excel"""

        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise HTTPException(
                status_code=500, detail="Excel export requires openpyxl package"
            )

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Company Comparison"

        # Header row with company names
        row_idx = 1
        worksheet.cell(row=row_idx, column=1, value="Metric")
        for col_idx, company_data in enumerate(companies, 2):
            cell = worksheet.cell(
                row=row_idx, column=col_idx, value=company_data.company_name_jp
            )
            cell.font = Font(bold=True)

        row_idx += 1

        # Ticker symbols
        worksheet.cell(row=row_idx, column=1, value="Ticker")
        for col_idx, company_data in enumerate(companies, 2):
            worksheet.cell(
                row=row_idx, column=col_idx, value=company_data.ticker_symbol
            )

        row_idx += 1
        worksheet.cell(row=row_idx, column=1, value="")  # Empty row
        row_idx += 1

        # Determine metrics to display
        metrics = request.metrics if request.metrics else []
        if not metrics and companies:
            # Extract all metrics from first company
            first_company = companies[0]
            for category in [
                "profitability",
                "safety",
                "efficiency",
                "growth",
                "valuation",
                "cash_flow",
            ]:
                category_data = getattr(first_company, category, {})
                if category_data:
                    metrics.extend(category_data.keys())

        # Write metrics rows
        for metric in metrics:
            label = ExportService._get_field_labels([metric])[0]
            worksheet.cell(row=row_idx, column=1, value=label)

            for col_idx, company_data in enumerate(companies, 2):
                # Find metric value in company data
                value = None
                for category in [
                    "profitability",
                    "safety",
                    "efficiency",
                    "growth",
                    "valuation",
                    "cash_flow",
                ]:
                    category_data = getattr(company_data, category, {})
                    if isinstance(category_data, dict) and metric in category_data:
                        value = category_data[metric]
                        break

                worksheet.cell(row=row_idx, column=col_idx, value=value)

                # Add ranking if requested
                if request.include_rankings and hasattr(company_data, "rankings"):
                    rankings = company_data.rankings or {}
                    if metric in rankings:
                        rank = rankings[metric]
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        if rank == 1:
                            cell.fill = PatternFill(
                                start_color="C6EFCE",
                                end_color="C6EFCE",
                                fill_type="solid",
                            )

            row_idx += 1

        # Summary section
        if summary:
            row_idx += 1
            worksheet.cell(row=row_idx, column=1, value="Summary Statistics").font = (
                Font(bold=True)
            )
            row_idx += 1

            worksheet.cell(row=row_idx, column=1, value="Best Overall")
            worksheet.cell(
                row=row_idx,
                column=2,
                value=getattr(summary, "best_overall", {}).get("company_name", "N/A"),
            )
            row_idx += 1

            worksheet.cell(row=row_idx, column=1, value="Total Companies")
            worksheet.cell(
                row=row_idx,
                column=2,
                value=getattr(summary, "total_companies", len(companies)),
            )

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = (
            request.filename or f"comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        )

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
        )

    @staticmethod
    def export_financial_data(
        db: Session, request: FinancialDataExportRequest
    ) -> StreamingResponse:
        """Export detailed financial data"""

        # Validate companies exist
        companies = db.query(Company).filter(Company.id.in_(request.company_ids)).all()

        if len(companies) != len(request.company_ids):
            raise HTTPException(
                status_code=404, detail="One or more companies not found"
            )

        # Only Excel format for financial data (multi-sheet support)
        if request.format == ExportFormat.EXCEL:
            return ExportService._export_financial_excel(db, companies, request)
        else:
            raise HTTPException(
                status_code=400,
                detail="Only Excel format is supported for financial data export",
            )

    @staticmethod
    def _export_financial_excel(
        db: Session, companies: List[Company], request: FinancialDataExportRequest
    ) -> StreamingResponse:
        """Export financial data as Excel with multiple sheets"""

        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font
        except ImportError:
            raise HTTPException(
                status_code=500, detail="Excel export requires openpyxl package"
            )

        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove default sheet

        # Create sheets for each company
        for company in companies:
            # Create sheet for this company
            sheet_name = f"{company.ticker_symbol}_{company.company_name_jp[:10]}"
            worksheet = workbook.create_sheet(title=sheet_name)

            row_idx = 1

            # Company header
            worksheet.cell(row=row_idx, column=1, value="Company")
            worksheet.cell(row=row_idx, column=2, value=company.company_name_jp)
            row_idx += 1
            worksheet.cell(row=row_idx, column=1, value="Ticker")
            worksheet.cell(row=row_idx, column=2, value=company.ticker_symbol)
            row_idx += 2

            # Financial Statements
            if "statements" in request.data_types:
                worksheet.cell(
                    row=row_idx, column=1, value="Financial Statements"
                ).font = Font(bold=True)
                row_idx += 1

                statements = (
                    db.query(FinancialStatement)
                    .filter(FinancialStatement.company_id == company.id)
                    .order_by(FinancialStatement.period_end.desc())
                    .limit(request.periods)
                    .all()
                )

                if statements:
                    # Headers
                    headers = [
                        "Period End",
                        "Revenue",
                        "Operating Income",
                        "Net Income",
                        "Total Assets",
                        "Equity",
                    ]
                    for col_idx, header in enumerate(headers, 1):
                        worksheet.cell(row=row_idx, column=col_idx, value=header)
                    row_idx += 1

                    # Data rows
                    for stmt in statements:
                        worksheet.cell(
                            row=row_idx,
                            column=1,
                            value=(
                                stmt.period_end.strftime("%Y-%m-%d")
                                if stmt.period_end
                                else ""
                            ),
                        )
                        worksheet.cell(row=row_idx, column=2, value=stmt.revenue)
                        worksheet.cell(
                            row=row_idx, column=3, value=stmt.operating_income
                        )
                        worksheet.cell(row=row_idx, column=4, value=stmt.net_income)
                        worksheet.cell(row=row_idx, column=5, value=stmt.total_assets)
                        worksheet.cell(
                            row=row_idx, column=6, value=stmt.shareholders_equity
                        )
                        row_idx += 1

                row_idx += 1

            # Financial Indicators
            if "indicators" in request.data_types:
                worksheet.cell(
                    row=row_idx, column=1, value="Financial Indicators"
                ).font = Font(bold=True)
                row_idx += 1

                indicators = (
                    db.query(FinancialIndicator)
                    .filter(FinancialIndicator.company_id == company.id)
                    .order_by(FinancialIndicator.date.desc())
                    .limit(request.periods)
                    .all()
                )

                if indicators:
                    # Headers
                    headers = [
                        "Date",
                        "ROE",
                        "ROA",
                        "Operating Margin",
                        "PER",
                        "PBR",
                        "Dividend Yield",
                    ]
                    for col_idx, header in enumerate(headers, 1):
                        worksheet.cell(row=row_idx, column=col_idx, value=header)
                    row_idx += 1

                    # Data rows
                    for ind in indicators:
                        worksheet.cell(
                            row=row_idx,
                            column=1,
                            value=ind.date.strftime("%Y-%m-%d") if ind.date else "",
                        )
                        worksheet.cell(row=row_idx, column=2, value=ind.roe)
                        worksheet.cell(row=row_idx, column=3, value=ind.roa)
                        worksheet.cell(
                            row=row_idx, column=4, value=ind.operating_margin
                        )
                        worksheet.cell(row=row_idx, column=5, value=ind.per)
                        worksheet.cell(row=row_idx, column=6, value=ind.pbr)
                        worksheet.cell(row=row_idx, column=7, value=ind.dividend_yield)
                        row_idx += 1

                row_idx += 1

            # Stock Prices
            if "stock_prices" in request.data_types:
                worksheet.cell(
                    row=row_idx, column=1, value="Stock Prices (Recent 30 days)"
                ).font = Font(bold=True)
                row_idx += 1

                stock_prices = (
                    db.query(StockPrice)
                    .filter(StockPrice.company_id == company.id)
                    .order_by(StockPrice.date.desc())
                    .limit(30)
                    .all()
                )

                if stock_prices:
                    # Headers
                    headers = ["Date", "Open", "High", "Low", "Close", "Volume"]
                    for col_idx, header in enumerate(headers, 1):
                        worksheet.cell(row=row_idx, column=col_idx, value=header)
                    row_idx += 1

                    # Data rows
                    for price in stock_prices:
                        worksheet.cell(
                            row=row_idx,
                            column=1,
                            value=price.date.strftime("%Y-%m-%d") if price.date else "",
                        )
                        worksheet.cell(row=row_idx, column=2, value=price.open_price)
                        worksheet.cell(row=row_idx, column=3, value=price.high_price)
                        worksheet.cell(row=row_idx, column=4, value=price.low_price)
                        worksheet.cell(row=row_idx, column=5, value=price.close_price)
                        worksheet.cell(row=row_idx, column=6, value=price.volume)
                        row_idx += 1

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 40)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = (
            request.filename
            or f"financial_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        )

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
        )
